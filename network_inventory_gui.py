import sys
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                           QHBoxLayout, QLabel, QLineEdit, QPushButton, 
                           QTextEdit, QProgressBar, QMessageBox, QTableWidget,
                           QTableWidgetItem, QFileDialog, QGroupBox, QTabWidget,
                           QCheckBox)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QRegularExpression, QSettings
from PyQt6.QtGui import QFont, QIcon, QRegularExpressionValidator
import getpass
import time
from netmiko import ConnectHandler
import re
from openpyxl import Workbook
from openpyxl.styles import Font

# Utility function for IP address validation
def get_ip_validator():
    """Returns a QRegularExpressionValidator for IP address validation"""
    ip_regex = QRegularExpression(
        "^(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$"
    )
    return QRegularExpressionValidator(ip_regex)

def validate_ip_address(ip_address):
    """Validates if a string is a valid IP address
    
    Args:
        ip_address (str): The IP address to validate
        
    Returns:
        bool: True if valid, False otherwise
    """
    ip_regex = QRegularExpression(
        "^(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$"
    )
    match = ip_regex.match(ip_address)
    return match.hasMatch()

class NetworkInventoryWorker(QThread):
    progress = pyqtSignal(str)
    finished = pyqtSignal(list)
    error = pyqtSignal(str)

    def __init__(self, username, password, distribution_ip):
        super().__init__()
        self.username = username
        self.password = password
        self.distribution_ip = distribution_ip
        self.device_type = "cisco_ios"
        self.discovered_switches = set()  # Track discovered switch IPs
        self.results = []
        self._abort = False

    def abort(self):
        self._abort = True

    def run(self):
        try:
            self.progress.emit(f"🔍 Starting network discovery from distribution switch {self.distribution_ip}")
            self.discover_network(self.distribution_ip, None)
            if self._abort:
                self.progress.emit("❌ Discovery cancelled by user.")
            self.finished.emit(self.results)
        except Exception as e:
            self.error.emit(str(e))

    def discover_network(self, ip, parent_ip):
        if self._abort:
            return
        if ip in self.discovered_switches:
            return

        self.discovered_switches.add(ip)
        self.progress.emit(f"🔌 Connecting to {ip}")
        
        switch_info = self.get_switch_info(ip, parent_ip)
        self.results.append(switch_info)

        # Get neighbors from CDP output
        device = {
            'device_type': self.device_type,
            'host': ip,
            'username': self.username,
            'password': self.password,
            'timeout': 20,
            'session_timeout': 30
        }

        try:
            with ConnectHandler(**device) as net_connect:
                full_cdp_output = self.get_full_output(net_connect, "show cdp neighbors detail")
                neighbors = re.findall(
                    r"Device ID: (\S+).*?IP address: (\d+\.\d+\.\d+\.\d+).*?Interface: (\S+),.*?Port ID \(outgoing port\): (\S+)",
                    full_cdp_output,
                    re.DOTALL
                )

                for device_id, neighbor_ip, local_intf, remote_intf in neighbors:
                    # Ignore CoreSW and GXP switches neighbors
                    if device_id.lower().startswith("core") or device_id.lower().startswith("gxp"):
                        continue
                    # Skip if neighbor is an AP or already discovered
                    if not self.is_ap_device(neighbor_ip) and neighbor_ip not in self.discovered_switches:
                        if self._abort:
                            return
                        self.progress.emit(f"🔍 Found new neighbor: {device_id} ({neighbor_ip})")
                        self.discover_network(neighbor_ip, ip)
                        time.sleep(1)  # Small delay between connections

        except Exception as e:
            self.progress.emit(f"❌ Failed to get neighbors for {ip}: {str(e)}")

    def get_switch_info(self, ip, parent_ip):
        device = {
            'device_type': self.device_type,
            'host': ip,
            'username': self.username,
            'password': self.password,
            'timeout': 20,
            'session_timeout': 30
        }

        try:
            with ConnectHandler(**device) as net_connect:
                hostname = net_connect.find_prompt().replace("#", "").strip()
                version_output = net_connect.send_command("show version", delay_factor=2)
                
                serial_match = re.search(r"(?i)System serial number\s*:\s*([A-Z0-9]+)", version_output)
                model_match = re.search(r"(?i)Model number\s*:\s*([\w+-]+)", version_output)
                model = model_match.group(1) if model_match else "Not Found"
                serial = serial_match.group(1) if serial_match else "Not Found"

                cdp_output = net_connect.send_command("show cdp neighbors detail", delay_factor=2)
                neighbors = re.findall(
                    r"Device ID: (\S+).*?IP address: (\d+\.\d+\.\d+\.\d+).*?Interface: (\S+),.*?Port ID \(outgoing port\): (\S+)",
                    cdp_output,
                    re.DOTALL
                )

                local_iface = "Unknown"
                uplink = "None"
                distribution_local_iface = "Unknown"
                distribution_uplink = "None"
                
                for device_id, neighbor_ip, local_intf, remote_intf in neighbors:
                    if parent_ip and neighbor_ip == parent_ip:
                        local_iface = local_intf
                        uplink = f"{neighbor_ip} ({remote_intf})"
                        break
                    if ip == self.distribution_ip and (device_id.lower().startswith("core") or device_id.lower().startswith("gxp")):
                        distribution_local_iface = local_intf
                        distribution_uplink = f"{neighbor_ip} ({remote_intf})"
                        break

                if ip == self.distribution_ip:
                    local_iface = distribution_local_iface
                    uplink = distribution_uplink

                extra_neighbors = []
                if parent_ip and ip != self.distribution_ip:
                    for device_id, neighbor_ip, local_intf, remote_intf in neighbors:
                        if (not self.is_ap_device(neighbor_ip)) and (neighbor_ip != parent_ip):
                            if neighbor_ip != uplink.split(' ')[0]:
                                extra_neighbors.append(f"{device_id} ({neighbor_ip}) via {local_intf} > {remote_intf}")

                return {
                    'name': hostname,
                    'ip': f"{ip} ({local_iface})",
                    'uplink': uplink,
                    'model': model,
                    'serial': serial,
                    'extra_neighbors': extra_neighbors if extra_neighbors else ["None"]
                }

        except Exception as e:
            self.progress.emit(f"❌ Failed {ip}: {str(e)}")
            return {
                'name': "Connection Failed",
                'ip': ip,
                'uplink': "N/A",
                'model': "N/A",
                'serial': str(e),
                'extra_neighbors': ["N/A"]
            }

    def get_full_output(self, net_connect, command):
        output = ""
        net_connect.send_command("\n")
        net_connect.write_channel(f"{command}\n")
        
        while True:
            time.sleep(1)
            chunk = net_connect.read_channel()
            output += chunk
            
            if "--More--" in chunk:
                net_connect.write_channel(" ")
            elif net_connect.find_prompt() in chunk:
                break
                
        return output

    def is_ap_device(self, neighbor_ip):
        return neighbor_ip.startswith("172.19.")

class NetworkDocumenterWidget(QWidget):
    def __init__(self):
        super().__init__()
        
        # Settings for "Remember Me" feature
        self.settings = QSettings("NetworkTools", "NetworkDocumenter")
        
        layout = QVBoxLayout(self)
        
        # Create credential group
        credential_group = QGroupBox("Login Credentials")
        credential_layout = QVBoxLayout()
        
        # Username input
        username_layout = QHBoxLayout()
        username_layout.addWidget(QLabel("Username:"))
        self.username_input = QLineEdit()
        username_layout.addWidget(self.username_input)
        credential_layout.addLayout(username_layout)
        
        # Password input
        password_layout = QHBoxLayout()
        password_layout.addWidget(QLabel("Password:"))
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        password_layout.addWidget(self.password_input)
        credential_layout.addLayout(password_layout)
        
        # Remember me checkbox
        self.remember_checkbox = QCheckBox("Remember Me")
        credential_layout.addWidget(self.remember_checkbox)
        
        credential_group.setLayout(credential_layout)
        layout.addWidget(credential_group)
        
        # Create buildings group
        self.buildings_group = QGroupBox("Buildings (Name + Distribution Switch IP)")
        self.buildings_layout = QVBoxLayout()
        self.building_rows = []
        self.add_building_row()  # Add initial row
        
        # Add button to add more buildings
        add_building_btn = QPushButton("+")
        add_building_btn.setFixedWidth(30)
        add_building_btn.clicked.connect(self.add_building_row)
        self.buildings_layout.addWidget(add_building_btn, alignment=Qt.AlignmentFlag.AlignLeft)
        
        self.buildings_group.setLayout(self.buildings_layout)
        layout.addWidget(self.buildings_group)
        
        # Create progress area
        self.progress_text = QTextEdit()
        self.progress_text.setReadOnly(True)
        layout.addWidget(self.progress_text)
        
        # Create progress bar
        self.progress_bar = QProgressBar()
        layout.addWidget(self.progress_bar)
        
        # Create buttons
        button_layout = QHBoxLayout()
        
        self.start_button = QPushButton("Start Network Discovery")
        self.start_button.clicked.connect(self.start_inventory)
        button_layout.addWidget(self.start_button)
        
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.cancel_inventory)
        self.cancel_button.setEnabled(False)
        button_layout.addWidget(self.cancel_button)
        
        self.export_button = QPushButton("Export to Excel")
        self.export_button.clicked.connect(self.export_to_excel)
        self.export_button.setEnabled(False)
        button_layout.addWidget(self.export_button)
        
        layout.addLayout(button_layout)
        
        # Create tab widget for results
        self.tab_widget = QTabWidget()
        layout.addWidget(self.tab_widget)
        
        self.results = []
        self.building_results = {}  # For multi-building
        self.building_tables = {}  # Store table widgets for each building
        
        # Load saved credentials if available
        self.load_saved_credentials()

    def load_saved_credentials(self):
        """Load saved credentials from settings"""
        # Check if "Remember Me" was previously checked
        remember_enabled = self.settings.value("remember_enabled", False, type=bool)
        
        if remember_enabled:
            # Load the saved username only
            self.username_input.setText(self.settings.value("username", ""))
            self.remember_checkbox.setChecked(True)
            
            # Don't load building data anymore

    def save_credentials(self):
        """Save credentials if Remember Me is checked"""
        if self.remember_checkbox.isChecked():
            self.settings.setValue("remember_enabled", True)
            self.settings.setValue("username", self.username_input.text())
            
            # Don't save building data anymore
        else:
            # Clear saved credentials if "Remember Me" is unchecked
            self.settings.setValue("remember_enabled", False)
            self.settings.remove("username")
    
    def add_building_row(self):
        row_widget = QWidget()
        row_layout = QHBoxLayout(row_widget)
        row_layout.setContentsMargins(0, 0, 0, 0)
        name_input = QLineEdit()
        name_input.setPlaceholderText("Building Name")
        ip_input = QLineEdit()
        ip_input.setPlaceholderText("Distribution Switch IP")
        
        # Add IP address validation
        ip_validator = get_ip_validator()
        # ip_input.setValidator(ip_validator) # Remove live validation
        
        row_layout.addWidget(QLabel("Name:"))
        row_layout.addWidget(name_input)
        row_layout.addWidget(QLabel("IP:"))
        row_layout.addWidget(ip_input)
        # Optionally, add a remove button for each row (except the first)
        if len(self.building_rows) > 0:
            remove_btn = QPushButton("-")
            remove_btn.setFixedWidth(30)
            def remove_row():
                self.buildings_layout.removeWidget(row_widget)
                row_widget.deleteLater()
                self.building_rows.remove((name_input, ip_input, row_widget))
            remove_btn.clicked.connect(remove_row)
            row_layout.addWidget(remove_btn)
        self.buildings_layout.insertWidget(len(self.building_rows), row_widget)
        self.building_rows.append((name_input, ip_input, row_widget))

    def create_results_table(self):
        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels([
            "Name", "IP (Interface)", "Uplink", 
            "Model Number", "Serial Number", "Extra Neighbors"
        ])
        return table

    def display_building_results(self, building_name):
        if building_name not in self.building_tables:
            # Create new table widget for this building
            table = self.create_results_table()
            self.building_tables[building_name] = table
            self.tab_widget.addTab(table, building_name)
        
        table = self.building_tables[building_name]
        results = self.building_results.get(building_name, [])
        table.setRowCount(len(results))
        
        for row, switch in enumerate(results):
            table.setItem(row, 0, QTableWidgetItem(switch['name']))
            table.setItem(row, 1, QTableWidgetItem(switch['ip']))
            table.setItem(row, 2, QTableWidgetItem(switch['uplink']))
            table.setItem(row, 3, QTableWidgetItem(switch['model']))
            table.setItem(row, 4, QTableWidgetItem(switch['serial']))
            table.setItem(row, 5, QTableWidgetItem('\n'.join(switch['extra_neighbors'])))
        
        table.resizeColumnsToContents()
        table.resizeRowsToContents()

    def start_inventory(self):
        username = self.username_input.text()
        password = self.password_input.text()
        building_inputs = [(row[0].text().strip(), row[1].text().strip()) for row in self.building_rows]
        
        if not username or not password:
            QMessageBox.warning(self, "Error", "Please enter both username and password")
            return
        
        # Filter out empty rows
        building_inputs = [(name, ip) for name, ip in building_inputs if name and ip]
        if not building_inputs:
            QMessageBox.warning(self, "Error", "Please enter at least one building name and distribution switch IP")
            return
        
        # Validate all IP addresses
        for building_name, ip in building_inputs:
            if not validate_ip_address(ip):
                QMessageBox.warning(self, "Error", f"Invalid IP address format for building '{building_name}': {ip}\nPlease use format: xxx.xxx.xxx.xxx")
                return
        
        # Save credentials
        self.save_credentials()
        
        self.progress_text.clear()
        self.progress_bar.setRange(0, 0)  # Indeterminate progress
        self.start_button.setEnabled(False)
        self.cancel_button.setEnabled(True)
        self.export_button.setEnabled(False)
        
        # Clear existing tabs and tables
        self.tab_widget.clear()
        self.building_tables.clear()
        self.building_results = {}
        
        self._current_building_idx = 0
        self._building_inputs = building_inputs
        self._username = username
        self._password = password
        self._run_next_building()

    def _run_next_building(self):
        if self._current_building_idx >= len(self._building_inputs):
            self.progress_bar.setRange(0, 100)
            self.progress_bar.setValue(100)
            self.start_button.setEnabled(True)
            self.cancel_button.setEnabled(False)
            self.export_button.setEnabled(True)
            self.progress_text.append("\n✅ All buildings completed.")
            return
        name, ip = self._building_inputs[self._current_building_idx]
        self.progress_text.append(f"\n=== Starting discovery for building: {name} ({ip}) ===")
        self.worker = NetworkInventoryWorker(self._username, self._password, ip)
        self.worker.progress.connect(self.update_progress)
        self.worker.finished.connect(lambda results, n=name: self._building_inventory_completed(n, results))
        self.worker.error.connect(self.handle_error)
        self.worker.start()

    def _building_inventory_completed(self, building_name, results):
        self.building_results[building_name] = results
        # Create and display the tab for this building immediately
        self.display_building_results(building_name)
        self._current_building_idx += 1
        self._run_next_building()

    def cancel_inventory(self):
        if hasattr(self, 'worker') and self.worker.isRunning():
            self.worker.abort()
            self.cancel_button.setEnabled(False)
            self.progress_text.append("⏹️ Cancelling discovery...")
        
    def update_progress(self, message):
        self.progress_text.append(message)
        
    def handle_error(self, error_message):
        QMessageBox.critical(self, "Error", f"An error occurred: {error_message}")
        self.start_button.setEnabled(True)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        
    def export_to_excel(self):
        if not self.building_results:
            QMessageBox.warning(self, "Error", "No results to export")
            return
        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            "",
            "Excel Files (*.xlsx)"
        )
        if not file_name:
            return
        try:
            wb = Workbook()
            # Remove the default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)
            for building_name, results in self.building_results.items():
                ws = wb.create_sheet(title=building_name[:31])  # Excel sheet name max length is 31
                headers = [
                    "Name", "IP (Interface)", "Uplink", 
                    "Model Number", "Serial Number", "Extra Neighbors"
                ]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                for row, switch in enumerate(results, 2):
                    ws.cell(row=row, column=1, value=switch['name'])
                    ws.cell(row=row, column=2, value=switch['ip'])
                    ws.cell(row=row, column=3, value=switch['uplink'])
                    ws.cell(row=row, column=4, value=switch['model'])
                    ws.cell(row=row, column=5, value=switch['serial'])
                    extra_neighbors = "\n".join(switch['extra_neighbors'])
                    ws.cell(row=row, column=6, value=extra_neighbors)
                    if len(switch['extra_neighbors']) > 1:
                        ws.row_dimensions[row].height = 15 * len(switch['extra_neighbors'])
                for col in ws.columns:
                    max_length = 0
                    for cell in col:
                        try:
                            if col[0].column_letter == 'F' and cell.value:
                                neighbor_lines = cell.value.split('\n')
                                max_line_length = max(len(line) for line in neighbor_lines)
                                if max_line_length > max_length:
                                    max_length = max_line_length
                            elif len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[col[0].column_letter].width = adjusted_width
            wb.save(file_name)
            QMessageBox.information(self, "Success", f"Inventory saved to '{file_name}'")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save Excel file: {str(e)}")

class MacTrackerWorker(QThread):
    progress = pyqtSignal(str)
    result = pyqtSignal(dict)
    error = pyqtSignal(str)
    finished_signal = pyqtSignal()  # Add a finished signal
    
    # Add log level constants
    LOG_INFO = 1    # Regular informational logs
    LOG_DEBUG = 2   # Detailed technical logs
    
    def __init__(self, username, password, distribution_ip, mac_suffix):
        super().__init__()
        self.username = username
        self.password = password
        self.distribution_ip = distribution_ip
        self.mac_suffix = mac_suffix.lower()  # Convert to lowercase for consistency
        self.device_type = "cisco_ios"
        self._abort = False
        self.active_connection = None  # Track the active connection
        self.log_level = self.LOG_INFO  # Default to showing only important logs
    
    def log(self, message, level=LOG_INFO):
        """Log messages with level filtering"""
        if level <= self.log_level:
            self.progress.emit(message)
            
    def set_log_level(self, level):
        """Set the logging level"""
        self.log_level = level

    def abort(self):
        self._abort = True
        # Disconnect if there's an active connection
        if self.active_connection:
            try:
                self.log("⏹️ Disconnecting active session...", self.LOG_INFO)
                self.active_connection.disconnect()
                self.active_connection = None
            except:
                pass

    def run(self):
        try:
            self.log(f"🔍 Starting MAC address search for suffix: {self.mac_suffix}", self.LOG_INFO)
            found = self.track_mac(self.distribution_ip)
            if self._abort:
                self.log("❌ Discovery cancelled by user.", self.LOG_INFO)
            elif not found:
                self.log(f"❌ MAC address with suffix {self.mac_suffix} was not found in the network.", self.LOG_INFO)
            self.finished_signal.emit()
        except Exception as e:
            self.error.emit(str(e))
            self.finished_signal.emit()

    def track_mac(self, switch_ip, depth=0):
        if self._abort:
            return False

        # Create indent based on search depth for better log readability
        indent = "  " * depth
        self.log(f"{indent}🔌 Connecting to switch {switch_ip}", self.LOG_INFO)

        device = {
            'device_type': self.device_type,
            'host': switch_ip,
            'username': self.username,
            'password': self.password,
            'timeout': 20,
            'session_timeout': 30
        }

        try:
            with ConnectHandler(**device) as net_connect:
                self.active_connection = net_connect
                # Get hostname
                hostname = net_connect.find_prompt().replace("#", "").strip()
                self.log(f"{indent}📡 Connected to {hostname}", self.LOG_INFO)
                
                # Search for our MAC suffix - don't log all the command details
                self.log(f"{indent}🔍 Searching for MAC suffix {self.mac_suffix}", self.LOG_INFO)
                
                # Try different commands and formats to find the MAC
                commands = [
                    f"show mac address-table",     # Standard command on newer IOS versions
                    f"show mac-address-table",     # Alternative format with hyphen for older IOS versions
                    f"show mac address-table dynamic",  # Explicitly shows only dynamic entries, sometimes needed on certain switches
                    f"show dot11 associations",
                ]
                
                mac_lines = []
                for command in commands:
                    if self._abort:
                        self.active_connection = None
                        return False
                        
                    # Don't log individual command execution
                    try:
                        mac_output = net_connect.send_command(command, delay_factor=2)
                        if mac_output and not "Invalid input" in mac_output:
                            # Process the output to find exact MAC suffix matches
                            for line in mac_output.strip().split('\n'):
                                # Only include lines that have the exact MAC suffix
                                # Look for the MAC suffix in various formats: 1234, 12:34, 12-34, etc.
                                if re.search(r'[^a-f0-9]' + re.escape(self.mac_suffix) + r'(?:[^a-f0-9]|$)', line.lower()):
                                    mac_lines.append(line)
                            
                            if mac_lines:
                                self.log(f"{indent}✅ Found MAC entries with suffix {self.mac_suffix}", self.LOG_INFO)
                                break
                    except Exception as e:
                        self.log(f"{indent}⚠️ Error searching MAC table: {str(e)}", self.LOG_DEBUG)
                
                if not mac_lines:
                    self.log(f"{indent}❌ MAC suffix {self.mac_suffix} not found on this switch.", self.LOG_INFO)
                    self.active_connection = None
                    return False
                
                found_result = False
                for mac_line in mac_lines:
                    if self._abort:
                        self.active_connection = None
                        return False
                    
                    # Debug level for specific entry details
                    self.log(f"{indent}🔎 Found MAC entry: {mac_line}", self.LOG_DEBUG)
                    
                    # Extract interface from the MAC table entry
                    interface = self._extract_interface_from_mac_line(mac_line)
                    
                    if not interface:
                        self.log(f"{indent}⚠️ Could not extract interface from MAC entry", self.LOG_DEBUG)
                        continue
                    
                    self.log(f"{indent}🔍 Checking interface {interface}", self.LOG_INFO)
                    
                    # Check if this interface is access or trunk
                    is_trunk = self._check_if_interface_is_trunk(net_connect, interface, indent)
                    
                    if is_trunk:
                        self.log(f"{indent}🔀 Interface is a trunk - checking next switch", self.LOG_INFO)
                        
                        # Get neighbor information
                        neighbor_info = self._get_neighbor_info(net_connect, interface, indent)
                        
                        if neighbor_info:
                            neighbor_ip, neighbor_name = neighbor_info
                            self.log(f"{indent}➡️ Following to switch: {neighbor_name}", self.LOG_INFO)
                            
                            # Before recursing, clear the active connection
                            self.active_connection = None
                            
                            # Recursively check the neighbor switch
                            found_deeper = self.track_mac(neighbor_ip, depth + 1)
                            if found_deeper:
                                return True
                        else:
                            self.log(f"{indent}⚠️ No CDP neighbor found on trunk", self.LOG_DEBUG)
                    else:
                        # Found MAC on access port - this is our final result
                        self.log(f"{indent}🎯 FOUND! Device with MAC suffix {self.mac_suffix} is on:", self.LOG_INFO)
                        self.log(f"{indent}   Switch: {hostname}", self.LOG_INFO)
                        self.log(f"{indent}   IP: {switch_ip}", self.LOG_INFO)
                        self.log(f"{indent}   Interface: {interface} (Access Port)", self.LOG_INFO)
                        
                        # Send the result
                        result = {
                            'hostname': hostname,
                            'ip': switch_ip,
                            'interface': interface,
                            'mac_suffix': self.mac_suffix,
                            'full_mac': self._extract_full_mac_from_line(mac_line)
                        }
                        self.result.emit(result)
                        found_result = True
                        self.active_connection = None
                        return True
                
                # Clear the active connection before returning
                self.active_connection = None
                return found_result
        
        except Exception as e:
            self.log(f"{indent}❌ Error with switch {switch_ip}: {str(e)}", self.LOG_INFO)
            self.error.emit(f"Failed while checking {switch_ip}: {str(e)}")
            self.active_connection = None
            return False
            
    def _extract_full_mac_from_line(self, mac_line):
        """Helper method to extract full MAC address from a MAC table line"""
        # Common MAC address formats: xxxx.xxxx.xxxx, xx:xx:xx:xx:xx:xx, xx-xx-xx-xx-xx-xx
        mac_match = re.search(r'([0-9a-fA-F]{4}\.[0-9a-fA-F]{4}\.[0-9a-fA-F]{4}|[0-9a-fA-F]{2}(?:[:-][0-9a-fA-F]{2}){5})', mac_line)
        if mac_match:
            return mac_match.group(0).lower() # Return in lowercase for consistency
        return "Not Found"

    def _extract_interface_from_mac_line(self, mac_line):
        """Helper method to extract interface from MAC table line"""
        interface = None
        
        # Pattern 1: Common format with interface at the end
        # e.g., "1    0050.56bf.beef    DYNAMIC    Gi1/0/1"
        interface_match = re.search(r'(?:(?:Gi|Fa|Te|Po|Et|Eth|Port-channel|GigabitEthernet|FastEthernet|TenGigabitEthernet)\S+)$', mac_line)
        if interface_match:
            interface = interface_match.group(0)
        
        # If not found, try splitting by whitespace and look for interface pattern
        if not interface:
            parts = mac_line.split()
            for part in parts:
                if re.match(r'^(?:Gi|Fa|Te|Po|Et|Eth|Port-channel|GigabitEthernet|FastEthernet|TenGigabitEthernet)\S+$', part):
                    interface = part
                    break
        
        # If still not found, assume it's the last column
        if not interface and len(mac_line.split()) > 0:
            interface = mac_line.split()[-1]
            
        return interface
        
    def _check_if_interface_is_trunk(self, net_connect, interface, indent):
        """Helper method to check if an interface is a trunk port"""
        switchport_output = net_connect.send_command(f"show interface {interface} switchport", delay_factor=2)
        
        # Check if it's a trunk port
        is_trunk = False
        admin_mode_match = re.search(r"Administrative Mode: (.+?)[\r\n]", switchport_output)
        oper_mode_match = re.search(r"Operational Mode: (.+?)[\r\n]", switchport_output)
        
        if admin_mode_match and "trunk" in admin_mode_match.group(1).lower():
            is_trunk = True
        if oper_mode_match and "trunk" in oper_mode_match.group(1).lower():
            is_trunk = True
        
        # Only log mode details at debug level
        admin_mode = admin_mode_match.group(1) if admin_mode_match else "unknown"
        oper_mode = oper_mode_match.group(1) if oper_mode_match else "unknown"
        self.log(f"{indent}🔄 Interface mode: {oper_mode}", self.LOG_DEBUG)
        
        return is_trunk
        
    def _get_neighbor_info(self, net_connect, interface, indent):
        """Helper method to get CDP neighbor information"""
        try:
            # Get CDP neighbor info for this interface
            cdp_output = net_connect.send_command(f"show cdp neighbor {interface} detail", delay_factor=2)
            
            # Extract neighbor IP and name
            ip_match = re.search(r"IP(?:\s?v4)?\s+[Aa]ddress(?:\s+or\s+subnet\s*)*: (\d+\.\d+\.\d+\.\d+)", cdp_output)
            device_id_match = re.search(r"Device ID:? (.+?)[\r\n]", cdp_output)
            
            if ip_match and device_id_match:
                neighbor_ip = ip_match.group(1)
                neighbor_name = device_id_match.group(1)
                return (neighbor_ip, neighbor_name)
            
            # If direct CDP lookup fails, try alternative methods
            self.log(f"{indent}⚠️ Trying alternative CDP lookup", self.LOG_DEBUG)
            
            # Try getting all CDP neighbors and filter for our interface
            all_cdp_output = net_connect.send_command("show cdp neighbors", delay_factor=2)
            
            # Try to find the interface in the CDP table
            cdp_lines = all_cdp_output.strip().split("\n")
            for line in cdp_lines:
                if interface in line:
                    # Found a matching line, extract device ID
                    parts = line.split()
                    if len(parts) >= 2:
                        neighbor_device = parts[0]
                        
                        # Get details for this neighbor
                        neighbor_detail = net_connect.send_command(f"show cdp entry {neighbor_device}", delay_factor=2)
                        ip_match = re.search(r"IP(?:\s?v4)?\s+[Aa]ddress(?:\s+or\s+subnet\s*)*: (\d+\.\d+\.\d+\.\d+)", neighbor_detail)
                        
                        if ip_match:
                            neighbor_ip = ip_match.group(1)
                            return (neighbor_ip, neighbor_device)
            
            return None
        except Exception as e:
            self.log(f"{indent}⚠️ Error getting CDP info: {str(e)}", self.LOG_DEBUG)
            return None

class MacTrackerWidget(QWidget):
    def __init__(self):
        super().__init__()
        
        # Settings for "Remember Me" feature
        self.settings = QSettings("NetworkTools", "MACTracker")
        
        # Main layout
        layout = QVBoxLayout(self)
        
        # Create credential group
        credential_group = QGroupBox("Login Credentials")
        credential_layout = QVBoxLayout()
        
        # Username input
        username_layout = QHBoxLayout()
        username_layout.addWidget(QLabel("Username:"))
        self.username_input = QLineEdit()
        username_layout.addWidget(self.username_input)
        credential_layout.addLayout(username_layout)
        
        # Password input
        password_layout = QHBoxLayout()
        password_layout.addWidget(QLabel("Password:"))
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        password_layout.addWidget(self.password_input)
        credential_layout.addLayout(password_layout)
        
        # Remember me checkbox
        self.remember_checkbox = QCheckBox("Remember Me")
        credential_layout.addWidget(self.remember_checkbox)
        
        credential_group.setLayout(credential_layout)
        layout.addWidget(credential_group)
        
        # Create search parameters group
        search_group = QGroupBox("MAC Search Parameters")
        search_layout = QVBoxLayout()
        
        # Distribution Switch IP input
        dist_ip_layout = QHBoxLayout()
        dist_ip_layout.addWidget(QLabel("Distribution Switch IP:"))
        self.dist_ip_input = QLineEdit()
        self.dist_ip_input.setPlaceholderText("e.g., 10.1.1.1")
        
        # Add IP address validation
        ip_validator = get_ip_validator()
        # self.dist_ip_input.setValidator(ip_validator) # Remove live validation
        
        dist_ip_layout.addWidget(self.dist_ip_input)
        search_layout.addLayout(dist_ip_layout)
        
        # Set initial focus to the IP input field
        self.dist_ip_input.setFocus()
        
        # MAC suffix input
        mac_suffix_layout = QHBoxLayout()
        mac_suffix_layout.addWidget(QLabel("Last 4 Digits of MAC Address:"))
        self.mac_suffix_input = QLineEdit()
        self.mac_suffix_input.setPlaceholderText("BEEF")  # Add placeholder text
        self.mac_suffix_input.setMaxLength(4)  # Limit to 4 characters
        
        # Use a validator instead of input mask for a more natural feel
        # Only allow hexadecimal input (0-9, a-f, A-F)
        hex_validator = QRegularExpressionValidator(QRegularExpression("[0-9A-Fa-f]{0,4}"))
        self.mac_suffix_input.setValidator(hex_validator)
        
        mac_suffix_layout.addWidget(self.mac_suffix_input)
        mac_suffix_help = QLabel("(Enter 4 hex digits: 0-9, A-F)")
        mac_suffix_help.setStyleSheet("color: gray;")
        mac_suffix_layout.addWidget(mac_suffix_help)
        search_layout.addLayout(mac_suffix_layout)
        
        # Add detailed logs checkbox
        self.detailed_logs_checkbox = QCheckBox("Show Detailed Technical Logs")
        self.detailed_logs_checkbox.setChecked(False)
        search_layout.addWidget(self.detailed_logs_checkbox)
        
        search_group.setLayout(search_layout)
        layout.addWidget(search_group)
        
        # Create buttons
        button_layout = QHBoxLayout()
        
        self.search_button = QPushButton("Track MAC")
        self.search_button.clicked.connect(self.start_tracking)
        button_layout.addWidget(self.search_button)
        
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.cancel_tracking)
        self.cancel_button.setEnabled(False)
        button_layout.addWidget(self.cancel_button)
        
        # Add clear logs button
        self.clear_button = QPushButton("Clear Logs")
        self.clear_button.clicked.connect(self.clear_logs)
        button_layout.addWidget(self.clear_button)
        
        layout.addLayout(button_layout)
        
        # Create progress area
        log_group = QGroupBox("Tracking Logs")
        log_layout = QVBoxLayout()
        
        self.progress_text = QTextEdit()
        self.progress_text.setReadOnly(True)
        log_layout.addWidget(self.progress_text)
        
        log_group.setLayout(log_layout)
        layout.addWidget(log_group)
        
        # Create results area
        results_group = QGroupBox("Results")
        results_layout = QVBoxLayout()
        
        self.result_text = QTextEdit()
        self.result_text.setReadOnly(True)
        results_layout.addWidget(self.result_text)
        
        results_group.setLayout(results_layout)
        layout.addWidget(results_group)
        
        # Load saved credentials if available
        self.load_saved_credentials()
    
    def load_saved_credentials(self):
        """Load saved credentials from settings"""
        # Check if "Remember Me" was previously checked
        remember_enabled = self.settings.value("remember_enabled", False, type=bool)
        
        if remember_enabled:
            # Load the saved username only
            self.username_input.setText(self.settings.value("username", ""))
            self.remember_checkbox.setChecked(True)
    
    def save_credentials(self):
        """Save credentials if Remember Me is checked"""
        if self.remember_checkbox.isChecked():
            self.settings.setValue("remember_enabled", True)
            self.settings.setValue("username", self.username_input.text())
            
            # Don't save building data anymore
        else:
            # Clear saved credentials if "Remember Me" is unchecked
            self.settings.setValue("remember_enabled", False)
            self.settings.remove("username")
    
    def clear_logs(self):
        """Clear the logs area"""
        self.progress_text.clear()
    
    def start_tracking(self):
        username = self.username_input.text()
        password = self.password_input.text()
        dist_ip = self.dist_ip_input.text()
        mac_suffix = self.mac_suffix_input.text()
        
        # Validate inputs
        if not username or not password:
            QMessageBox.warning(self, "Error", "Please enter both username and password")
            return
        
        if not dist_ip:
            QMessageBox.warning(self, "Error", "Please enter distribution switch IP")
            return
        
        # Additional IP validation to ensure it's in correct format
        if not validate_ip_address(dist_ip):
            QMessageBox.warning(self, "Error", "Please enter a valid IP address in format: xxx.xxx.xxx.xxx")
            return
        
        # Validate MAC suffix
        if not mac_suffix or len(mac_suffix) != 4 or not re.match(r'^[0-9a-fA-F]{4}$', mac_suffix):
            QMessageBox.warning(self, "Error", "Please enter exactly 4 valid hexadecimal characters (0-9, A-F) for MAC suffix")
            return
        
        # Save credentials if Remember Me is checked
        self.save_credentials()
        
        # Convert MAC suffix to lowercase for consistent matching
        mac_suffix = mac_suffix.lower()
        
        # Clear previous results
        self.progress_text.clear()
        self.result_text.clear()
        
        # Update UI state
        self.search_button.setEnabled(False)
        self.cancel_button.setEnabled(True)
        
        # Start worker thread
        self.worker = MacTrackerWorker(username, password, dist_ip, mac_suffix)
        
        # Set log level based on checkbox
        if self.detailed_logs_checkbox.isChecked():
            self.worker.set_log_level(MacTrackerWorker.LOG_DEBUG)
        else:
            self.worker.set_log_level(MacTrackerWorker.LOG_INFO)
            
        self.worker.progress.connect(self.update_progress)
        self.worker.result.connect(self.display_result)
        self.worker.error.connect(self.handle_error)
        self.worker.finished_signal.connect(self.finished_tracking)
        self.worker.start()
    
    def cancel_tracking(self):
        if hasattr(self, 'worker') and self.worker.isRunning():
            self.progress_text.append("⏹️ Cancelling MAC tracking...")
            self.worker.abort()
            self.cancel_button.setEnabled(False)
            self.search_button.setEnabled(True)
    
    def update_progress(self, message):
        self.progress_text.append(message)
        # Auto-scroll to the bottom
        scrollbar = self.progress_text.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())
    
    def display_result(self, result):
        # Format the final result in a nice way
        formatted_result = f"""
<h3>🎯 Device Found!</h3>
<table>
  <tr>
    <td><b>MAC Address:</b></td>
    <td>{result['full_mac']}</td>
  </tr>
  <tr>
    <td><b>Switch Name:</b></td>
    <td>{result['hostname']}</td>
  </tr>
  <tr>
    <td><b>Switch IP:</b></td>
    <td>{result['ip']}</td>
  </tr>
  <tr>
    <td><b>Interface:</b></td>
    <td>{result['interface']}</td>
  </tr>
</table>
<p>✅ The device is directly connected to this switch.</p>
"""
        self.result_text.setHtml(formatted_result)
        
        # Reset UI state
        self.search_button.setEnabled(True)
        self.cancel_button.setEnabled(False)
    
    def handle_error(self, error_message):
        QMessageBox.critical(self, "Error", f"An error occurred: {error_message}")
        self.search_button.setEnabled(True)
        self.cancel_button.setEnabled(False)
    
    def finished_tracking(self):
        self.progress_text.append("\n✅ MAC tracking completed.")
        self.search_button.setEnabled(True)
        self.cancel_button.setEnabled(False)

class NetworkInventoryGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Network Tools")
        self.setMinimumSize(800, 600)
        
        # Create main widget with tab widget
        self.main_tab_widget = QTabWidget()
        self.setCentralWidget(self.main_tab_widget)
        
        # Create Network Documenter tab
        self.network_documenter = NetworkDocumenterWidget()
        self.main_tab_widget.addTab(self.network_documenter, "Network Documenter")
        
        # Create MAC Tracker tab
        self.mac_tracker = MacTrackerWidget()
        self.main_tab_widget.addTab(self.mac_tracker, "MAC Tracker")
        
        # Future tabs can be added here
        # Example:
        # self.some_other_feature = SomeOtherFeatureWidget()
        # self.main_tab_widget.addTab(self.some_other_feature, "Some Other Feature")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = NetworkInventoryGUI()
    window.show()
    sys.exit(app.exec()) 